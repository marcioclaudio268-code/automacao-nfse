import csv
import json
import os
import re
import subprocess
import traceback
import sys
import unicodedata
from datetime import datetime
from pathlib import Path

from application.artifact_locator_service import ArtifactLocatorService
from core.config_runtime import competencia_alvo_dir_name
from core.app_info import APP_MAIN_BUNDLE_NAME, APP_MAIN_EXE_NAME
from core.company_paths import nome_pasta_empresa_por_dados, normalizar_codigo_empresa
from core.paths import build_runtime_paths

try:
    from openpyxl import Workbook

    RESUMO_USA_XLSX = True
except Exception:
    Workbook = None
    RESUMO_USA_XLSX = False


def get_runtime_base() -> Path:
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent
    return Path(__file__).resolve().parent


def get_main_command() -> list[str]:
    base = get_runtime_base()

    if getattr(sys, "frozen", False):
        main_exe = base / APP_MAIN_BUNDLE_NAME / APP_MAIN_EXE_NAME
        if not main_exe.exists():
            raise FileNotFoundError(f"Executavel do backend principal nao encontrado: {main_exe}")
        return [str(main_exe)]

    main_py = base / "main.py"
    if not main_py.exists():
        raise FileNotFoundError(f"Arquivo main.py nao encontrado: {main_py}")

    return [sys.executable, str(main_py)]


def executar_main_processo(env_extra: dict | None = None, timeout: int | None = None) -> subprocess.CompletedProcess:
    env = os.environ.copy()
    if env_extra:
        env.update(env_extra)

    cmd = get_main_command()
    return subprocess.run(
        cmd,
        env=env,
        cwd=str(get_runtime_base()),
        timeout=timeout,
    )


BASE_DIR = str(get_runtime_base())
OUTPUT_BASE_DIR = os.environ.get("OUTPUT_BASE_DIR", BASE_DIR)
os.makedirs(OUTPUT_BASE_DIR, exist_ok=True)

CSV_EMPRESAS = os.environ.get("EMPRESAS_ARQUIVO", os.environ.get("EMPRESAS_CSV", "empresas.xlsx"))
REPORT_PATH = os.path.join(OUTPUT_BASE_DIR, "report_execucao_empresas.csv")
CHECKPOINT_PATH = os.path.join(OUTPUT_BASE_DIR, "checkpoint_execucao_empresas.json")
RESUMO_BASE_NAME = "resumo_execucao_empresas"
RESUMO_HEADER = [
    "indice_lista",
    "linha_planilha",
    "empresa",
    "codigo",
    "cnpj",
    "competencia",
    "status_execucao",
    "teve_iss",
    "teve_prestados",
    "teve_tomados",
    "erro_tipo",
    "erro_resumo",
]
ARTIFACT_LOCATOR = ArtifactLocatorService()
MAX_TENTATIVAS = int(os.environ.get("MAX_TENTATIVAS_EMPRESA", "3"))
LOGIN_WAIT_SECONDS = int(os.environ.get("LOGIN_WAIT_SECONDS", "120"))
TIMEOUT_PROCESSO_MAIN = int(os.environ.get("TIMEOUT_PROCESSO_MAIN", "1800"))
CONTINUAR_DE_ONDE_PAROU = os.environ.get("CONTINUAR_DE_ONDE_PAROU", "1").strip() == "1"
USAR_CHECKPOINT = os.environ.get("USAR_CHECKPOINT", "1").strip() == "1"
MSG_CAPTCHA_TIMEOUT = "CAPTCHA_NAO_RESOLVIDO_NO_TEMPO"
EXIT_CODE_CAPTCHA_TIMEOUT = 30
EXIT_CODE_SEM_COMPETENCIA = 40
EXIT_CODE_SEM_SERVICOS = 41
EXIT_CODE_CREDENCIAL_INVALIDA = 50
EXIT_CODE_TOMADOS_PDF_REVISAO_MANUAL = 51
EXIT_CODE_EMPRESA_MULTIPLA = 52
EXIT_CODE_APURACAO_COMPLETA_REVISAO_MANUAL = 53
EXIT_CODE_MULTI_CADASTRO_REVISAO_MANUAL = 54
EXIT_CODE_TOMADOS_FALHA = 42
EXIT_CODE_CHROME_INIT_FALHA = 60
ARQUIVO_STATUS_TOMADOS_PDF = "_tomados_pdf_status.json"
STATUS_SUCESSO = {"SUCESSO", "SUCESSO_SEM_COMPETENCIA", "SUCESSO_SEM_SERVICOS"}
STATUS_FINAIS_NAO_REPROCESSAR = STATUS_SUCESSO | {"REVISAO_MANUAL"}
ERRO_TIPOS_OPERACIONAIS = {
    "LOGIN_INVALIDO",
    "ARQUIVO",
    "CAPTCHA",
    "TIMEOUT",
    "ERRO_PORTAL",
    "SEM_MOVIMENTO",
    "DESCONHECIDO",
}
ERRO_TIPOS_LEGADOS = {
    "CREDENCIAL_INVALIDA": "LOGIN_INVALIDO",
    "CAPTCHA_TIMEOUT": "CAPTCHA",
    "CAPTCHA_INCORRETO": "CAPTCHA",
    "TOMADOS_PDF": "ARQUIVO",
    "TOMADOS_FALHA": "ARQUIVO",
    "CHROME_INIT_FALHA": "ERRO_PORTAL",
    "EMPRESA_MULTIPLA": "ERRO_PORTAL",
    "MULTI_CADASTRO": "ERRO_PORTAL",
    "APURACAO_COMPLETA": "ERRO_PORTAL",
    "REVISAO_MANUAL": "ERRO_PORTAL",
    "SEM_COMPETENCIA": "SEM_MOVIMENTO",
    "SEM_SERVICOS": "SEM_MOVIMENTO",
    "SUCESSO_SEM_COMPETENCIA": "SEM_MOVIMENTO",
    "SUCESSO_SEM_SERVICOS": "SEM_MOVIMENTO",
}


def inferir_acao_recomendada(status: str, motivo: str = "") -> str:
    status_limpo = (status or "").strip().upper()
    motivo_limpo = (motivo or "").strip().lower()

    if status_limpo == "REVISAO_MANUAL":
        if "tomados" in motivo_limpo and "pdf" in motivo_limpo:
            return "ABRIR_DEBUG"
        return "CONFERIR_SENHA"
    if status_limpo != "FALHA":
        return ""
    if "captcha" in motivo_limpo:
        return "REPROCESSAR"
    return "ABRIR_DEBUG"


def caminho_status_tomados_pdf_empresa(empresa: dict) -> str:
    pasta_empresa = Path(OUTPUT_BASE_DIR) / "downloads" / nome_pasta_empresa_por_dados(empresa)
    legado = pasta_empresa / ARQUIVO_STATUS_TOMADOS_PDF
    if legado.exists():
        return str(legado)

    competencias = []
    if pasta_empresa.exists():
        for child in pasta_empresa.iterdir():
            if child.is_dir() and re.fullmatch(r"\d{2}\.\d{4}", child.name):
                competencias.append(child)

    competencias.sort(key=lambda path: (int(path.name.split(".")[1]), int(path.name.split(".")[0])), reverse=True)
    for competencia_dir in competencias:
        geral_dir = competencia_dir / "_GERAL"
        candidato = geral_dir / ARQUIVO_STATUS_TOMADOS_PDF
        if candidato.exists():
            return str(candidato)
        prefixed = sorted(
            geral_dir.glob(f"*{ARQUIVO_STATUS_TOMADOS_PDF}"),
            key=lambda path: path.stat().st_mtime,
            reverse=True,
        )
        if prefixed:
            return str(prefixed[0])

    return str(legado)


def carregar_status_tomados_pdf_empresa(empresa: dict) -> dict:
    path = caminho_status_tomados_pdf_empresa(empresa)
    if not os.path.exists(path):
        return {}

    try:
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
        return data if isinstance(data, dict) else {}
    except Exception:
        return {}


def normalizar_header(h: str) -> str:
    txt = (h or "").strip().lower()
    txt = unicodedata.normalize("NFKD", txt)
    return "".join(ch for ch in txt if not unicodedata.combining(ch))


def mapear_colunas(headers_normalizados):
    col_codigo = None
    col_razao = None
    col_cnpj = None
    col_segmento = None
    col_senha = None

    for h in headers_normalizados:
        if h.startswith("cod"):
            col_codigo = h
        elif "razao" in h:
            col_razao = h
        elif h == "cnpj":
            col_cnpj = h
        elif "segmento" in h:
            col_segmento = h
        elif "senha" in h:
            col_senha = h

    if not all([col_codigo, col_razao, col_cnpj, col_segmento, col_senha]):
        raise ValueError(
            "Colunas obrigatÃ³rias nÃ£o encontradas. Esperado: CÃ³digo, RazÃ£o Social, CNPJ, Segmento, Senha Prefeitura"
        )

    return col_codigo, col_razao, col_cnpj, col_segmento, col_senha


def carregar_empresas_csv(path_csv: str):
    empresas = []
    with open(path_csv, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f, delimiter=';')
        headers_normalizados = [normalizar_header(h) for h in (reader.fieldnames or [])]
        col_codigo, col_razao, col_cnpj, col_segmento, col_senha = mapear_colunas(headers_normalizados)

        for linha_planilha, raw in enumerate(reader, start=2):
            row = {normalizar_header(k): (v or "").strip() for k, v in raw.items()}
            if not row.get(col_cnpj):
                continue
            empresas.append({
                "codigo": row.get(col_codigo, ""),
                "razao_social": row.get(col_razao, ""),
                "cnpj": row.get(col_cnpj, ""),
                "segmento": row.get(col_segmento, ""),
                "senha_prefeitura": row.get(col_senha, ""),
                "indice_lista": len(empresas) + 1,
                "linha_planilha": linha_planilha,
            })
    return empresas


def encontrar_header_xlsx(ws, max_linhas_busca=25):
    linhas = ws.iter_rows(values_only=True)

    for i, row in enumerate(linhas, start=1):
        if i > max_linhas_busca:
            break

        header_norm = [normalizar_header(str(h) if h is not None else "") for h in row]
        try:
            cols = mapear_colunas(header_norm)
            idx = {h: j for j, h in enumerate(header_norm)}
            return i, cols, idx
        except ValueError:
            continue

    raise ValueError(
        "Colunas obrigatÃ³rias nÃ£o encontradas nas primeiras linhas do XLSX. "
        "Esperado: CÃ³digo, RazÃ£o Social, CNPJ, Segmento, Senha Prefeitura"
    )


def carregar_empresas_xlsx(path_xlsx: str):
    try:
        from openpyxl import load_workbook
    except Exception as e:
        raise RuntimeError("Para usar .xlsx instale a dependÃªncia: pip install openpyxl") from e

    wb = load_workbook(path_xlsx, read_only=True, data_only=True)
    ws = wb.active

    linha_header, (col_codigo, col_razao, col_cnpj, col_segmento, col_senha), idx = encontrar_header_xlsx(ws)

    empresas = []
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if i <= linha_header:
            continue

        vals = ["" if v is None else str(v).strip() for v in row]
        cnpj = vals[idx[col_cnpj]] if idx[col_cnpj] < len(vals) else ""
        if not cnpj:
            continue

        empresas.append({
            "codigo": vals[idx[col_codigo]] if idx[col_codigo] < len(vals) else "",
            "razao_social": vals[idx[col_razao]] if idx[col_razao] < len(vals) else "",
            "cnpj": cnpj,
            "segmento": vals[idx[col_segmento]] if idx[col_segmento] < len(vals) else "",
            "senha_prefeitura": vals[idx[col_senha]] if idx[col_senha] < len(vals) else "",
            "indice_lista": len(empresas) + 1,
            "linha_planilha": i,
        })

    return empresas


def carregar_empresas(path_arquivo: str):
    ext = os.path.splitext(path_arquivo)[1].lower()
    if ext == ".xlsx":
        return carregar_empresas_xlsx(path_arquivo)
    return carregar_empresas_csv(path_arquivo)


def filtrar_empresas_para_execucao(empresas, inicio=None, fim=None):
    if inicio is None or fim is None:
        return empresas
    return [
        empresa
        for empresa in empresas
        if inicio <= int(empresa.get("indice_lista", 0)) <= fim
    ]


def resolver_faixa_execucao(empresas):
    raw_inicio = os.environ.get("EMPRESA_INICIO", "").strip()
    raw_fim = os.environ.get("EMPRESA_FIM", "").strip()

    if not raw_inicio and not raw_fim:
        return None, None, empresas

    if not raw_inicio or not raw_fim:
        raise ValueError(
            "Faixa de execucao invalida. Informe EMPRESA_INICIO e EMPRESA_FIM, ou deixe ambos vazios."
        )

    try:
        inicio = int(raw_inicio)
        fim = int(raw_fim)
    except Exception as e:
        raise ValueError(
            "Faixa de execucao invalida. EMPRESA_INICIO e EMPRESA_FIM devem ser inteiros positivos."
        ) from e

    if inicio <= 0 or fim <= 0:
        raise ValueError(
            "Faixa de execucao invalida. EMPRESA_INICIO e EMPRESA_FIM devem ser inteiros positivos."
        )
    if inicio > fim:
        raise ValueError("Faixa de execucao invalida. EMPRESA_INICIO nao pode ser maior que EMPRESA_FIM.")

    total = len(empresas)
    if inicio > total:
        raise ValueError(
            f"Faixa de execucao invalida. EMPRESA_INICIO={inicio} excede o total carregado ({total})."
        )

    fim = min(fim, total)
    empresas_filtradas = filtrar_empresas_para_execucao(empresas, inicio, fim)
    return inicio, fim, empresas_filtradas


def _mapear_erro_por_codigo(registros_fonte: list[dict]) -> dict[str, str]:
    mapa: dict[str, str] = {}
    for registro in registros_fonte or []:
        empresa = registro.get("empresa") or {}
        resultado = registro.get("resultado") or {}
        codigo = normalizar_codigo_empresa(empresa.get("codigo") or empresa.get("codigo_empresa") or "")
        if not codigo:
            continue
        mapa[codigo] = classificar_erro_execucao(
            resultado.get("status") or resultado.get("status_execucao") or "",
            resultado.get("motivo") or resultado.get("erro_resumo") or "",
            resultado.get("erro_tipo") or "",
        )
    return mapa


def filtrar_empresas_por_criterios(
    empresas: list[dict],
    empresas_explicitadas: tuple[str, ...] = (),
    tipos_erro: tuple[str, ...] = (),
    registros_fonte: list[dict] | None = None,
) -> list[dict]:
    selecionadas = list(empresas)

    if empresas_explicitadas:
        codigos_explicitados = {normalizar_codigo_empresa(codigo) for codigo in empresas_explicitadas if codigo}
        selecionadas = [
            empresa
            for empresa in selecionadas
            if normalizar_codigo_empresa(empresa.get("codigo", "")) in codigos_explicitados
        ]

    if tipos_erro:
        if registros_fonte is None:
            raise ValueError(
                "FILTRAR_ERRO_TIPO foi informado, mas nenhuma fonte de resumo/report foi encontrada para o escopo atual."
            )
        categorias_desejadas = {categoria for categoria in tipos_erro if categoria}
        mapa_erro = _mapear_erro_por_codigo(registros_fonte)
        selecionadas = [
            empresa
            for empresa in selecionadas
            if mapa_erro.get(normalizar_codigo_empresa(empresa.get("codigo", "")), "DESCONHECIDO")
            in categorias_desejadas
        ]

    return selecionadas


def resolver_paths_execucao(output_base_dir, inicio=None, fim=None):
    report_path = os.path.join(output_base_dir, "report_execucao_empresas.csv")
    checkpoint_path = os.path.join(output_base_dir, "checkpoint_execucao_empresas.json")

    if inicio is None or fim is None:
        return report_path, checkpoint_path

    sufixo = _sufixo_lote(inicio, fim)
    report_path = os.path.join(output_base_dir, f"report_execucao_empresas{sufixo}.csv")
    checkpoint_path = os.path.join(output_base_dir, f"checkpoint_execucao_empresas{sufixo}.json")
    return report_path, checkpoint_path


def _sufixo_lote(inicio=None, fim=None) -> str:
    if inicio is None or fim is None:
        return ""

    largura = max(3, len(str(max(int(inicio), int(fim)))))
    return f"__lote_{int(inicio):0{largura}d}_{int(fim):0{largura}d}"


def resolver_path_resumo_execucao(output_base_dir, inicio=None, fim=None) -> str:
    ext = ".xlsx" if RESUMO_USA_XLSX else ".csv"
    sufixo = _sufixo_lote(inicio, fim)
    return os.path.join(output_base_dir, f"{RESUMO_BASE_NAME}{sufixo}{ext}")


def _normalizar_token_erro(valor: str) -> str:
    token = (valor or "").strip().upper()
    token = token.replace("-", "_").replace(" ", "_")
    token = re.sub(r"[^A-Z0-9_]+", "_", token)
    token = re.sub(r"_+", "_", token).strip("_")
    return token


def _resolver_categoria_erro_entrada(valor: str) -> str:
    token = _normalizar_token_erro(valor)
    if not token:
        return ""

    token = ERRO_TIPOS_LEGADOS.get(token, token)
    if token in ERRO_TIPOS_OPERACIONAIS:
        return token
    raise ValueError(
        f"FILTRAR_ERRO_TIPO invalido: {valor!r}. Valores aceitos: {', '.join(sorted(ERRO_TIPOS_OPERACIONAIS))}."
    )


def resolver_filtros_execucao() -> dict:
    raw_erro = os.environ.get("FILTRAR_ERRO_TIPO", "").strip()
    raw_empresas = os.environ.get("EMPRESAS", "").strip()

    tipos_erro: list[str] = []
    if raw_erro:
        for token in re.split(r"[,\s;|]+", raw_erro):
            if not token.strip():
                continue
            categoria = _resolver_categoria_erro_entrada(token)
            if categoria and categoria not in tipos_erro:
                tipos_erro.append(categoria)
        if not tipos_erro:
            raise ValueError(
                "FILTRAR_ERRO_TIPO foi informado, mas nenhum valor valido foi encontrado."
            )

    empresas: list[str] = []
    if raw_empresas:
        for token in re.split(r"[,\s;|]+", raw_empresas):
            if not token.strip():
                continue
            codigo = normalizar_codigo_empresa(token)
            if codigo and codigo not in empresas:
                empresas.append(codigo)
        if not empresas:
            raise ValueError("EMPRESAS foi informado, mas nenhum codigo valido foi encontrado.")

    return {
        "raw_erro": raw_erro,
        "raw_empresas": raw_empresas,
        "erro_tipos": tuple(tipos_erro),
        "empresas": tuple(empresas),
    }


def _contem_qualquer(texto: str, marcadores: tuple[str, ...]) -> bool:
    return any(marcador in texto for marcador in marcadores)


def classificar_erro_execucao(status: str, motivo: str = "", erro_tipo_previo: str = "") -> str:
    status_norm = _normalizar_token_erro(status)
    motivo_norm = _resumir_texto(motivo, 500).lower()
    erro_tipo_prev_raw = _normalizar_token_erro(erro_tipo_previo)
    erro_tipo_norm = ERRO_TIPOS_LEGADOS.get(erro_tipo_prev_raw, erro_tipo_prev_raw)
    if erro_tipo_prev_raw not in {"", "REVISAO_MANUAL", "DESCONHECIDO"} and erro_tipo_norm in ERRO_TIPOS_OPERACIONAIS:
        return erro_tipo_norm
    if erro_tipo_prev_raw in {"ERRO_PORTAL", "DESCONHECIDO"}:
        return erro_tipo_norm

    if status_norm in {"SUCESSO_SEM_COMPETENCIA", "SUCESSO_SEM_SERVICOS", "SEM_COMPETENCIA", "SEM_SERVICOS"}:
        return "SEM_MOVIMENTO"
    if status_norm == "SUCESSO":
        if _contem_qualquer(
            motivo_norm,
            (
                "sem notas na competencia",
                "sem modulo de nota fiscal",
                "sem movimento",
            ),
        ):
            return "SEM_MOVIMENTO"
        return ""

    if _contem_qualquer(
        motivo_norm,
        (
            "sem notas na competencia",
            "sem modulo de nota fiscal",
            "sem movimento",
        ),
    ):
        return "SEM_MOVIMENTO"
    if _contem_qualquer(
        motivo_norm,
        (
            "captcha",
            "captcha_n",
        ),
    ):
        return "CAPTCHA"
    if _contem_qualquer(
        motivo_norm,
        (
            "credencial invalida",
            "login invalido",
            "senha invalida",
            "senha incorreta",
            "usuario invalido",
            "login incorreto",
        ),
    ):
        return "LOGIN_INVALIDO"
    if _contem_qualquer(
        motivo_norm,
        (
            "arquivo",
            "xml",
            "pdf",
            "download",
            "crdownload",
            "nao gerado",
            "sem location",
            "resposta invalida",
            "file not found",
            "not found",
            "servicos tomados",
            "tomados pdf",
            "pdf tomados",
        ),
    ):
        return "ARQUIVO"
    if _contem_qualquer(
        motivo_norm,
        (
            "timeout",
            "temporiz",
        ),
    ):
        return "TIMEOUT"
    if _contem_qualquer(
        motivo_norm,
        (
            "empresa multipla",
            "multiplos cadastros",
            "apuracao completa",
            "webdriver",
            "chrome",
            "portal",
            "browser",
            "naveg",
            "sessao",
            "http 502",
            "http 503",
            " 502",
            " 503",
            "falta de sessao",
            "sessao expirada",
            "falha inicializar chrome",
            "falha na inicializacao do chrome",
        ),
    ):
        return "ERRO_PORTAL"
    if status_norm == "REVISAO_MANUAL":
        return "ERRO_PORTAL" if _contem_qualquer(
            motivo_norm,
            (
                "empresa multipla",
                "multiplos cadastros",
                "apuracao completa",
                "webdriver",
                "chrome",
                "portal",
                "browser",
                "naveg",
                "sessao",
                "http 502",
                "http 503",
                " 502",
                " 503",
                "falha inicializar chrome",
                "falha na inicializacao do chrome",
            ),
        ) else "DESCONHECIDO"
    if status_norm == "FALHA":
        return "ERRO_PORTAL" if _contem_qualquer(
            motivo_norm,
            (
                "empresa multipla",
                "multiplos cadastros",
                "apuracao completa",
                "webdriver",
                "chrome",
                "portal",
                "browser",
                "naveg",
                "sessao",
                "http 502",
                "http 503",
                " 502",
                " 503",
                "falha inicializar chrome",
                "falha na inicializacao do chrome",
            ),
        ) else "DESCONHECIDO"

    return "DESCONHECIDO"


def carregar_report_existente(path_report: str):
    concluidas = set()
    if not os.path.exists(path_report):
        return concluidas

    try:
        with open(path_report, "r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f, delimiter=';')
            for row in reader:
                status = (row.get("status") or "").strip().upper()
                cnpj = re.sub(r"\D", "", row.get("cnpj") or "")
                if status in STATUS_FINAIS_NAO_REPROCESSAR and cnpj:
                    concluidas.add(cnpj)
    except Exception:
        pass

    return concluidas


def carregar_rows_report_existente(path_report: str):
    rows = []
    if not os.path.exists(path_report):
        return rows

    try:
        with open(path_report, "r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f, delimiter=';')
            for row in reader:
                ts_inicio_raw = (row.get("timestamp_inicio") or "").strip()
                ts_fim_raw = (row.get("timestamp_fim") or "").strip()
                try:
                    ts_inicio = datetime.strptime(ts_inicio_raw, "%Y-%m-%d %H:%M:%S")
                except Exception:
                    ts_inicio = datetime.now()
                try:
                    ts_fim = datetime.strptime(ts_fim_raw, "%Y-%m-%d %H:%M:%S")
                except Exception:
                    ts_fim = ts_inicio

                tentativas_raw = (row.get("tentativas") or "0").strip()
                try:
                    tentativas = int(tentativas_raw or 0)
                except Exception:
                    tentativas = 0

                empresa = {
                    "codigo": (row.get("codigo_empresa") or "").strip(),
                    "razao_social": (row.get("razao_social") or "").strip(),
                    "cnpj": (row.get("cnpj") or "").strip(),
                    "segmento": (row.get("segmento") or "").strip(),
                }
                resultado = {
                    "status": (row.get("status") or "").strip(),
                    "motivo": (row.get("motivo") or "").strip(),
                    "tentativas": tentativas,
                    "acao_recomendada": (row.get("acao_recomendada") or "").strip(),
                }
                if not resultado["acao_recomendada"]:
                    resultado["acao_recomendada"] = inferir_acao_recomendada(
                        resultado["status"],
                        resultado["motivo"],
                    )
                rows.append({"empresa": empresa, "resultado": resultado, "inicio": ts_inicio, "fim": ts_fim})
    except Exception:
        return rows

    return rows


def _texto_campo(valor) -> str:
    return "" if valor is None else str(valor).strip()


def _registro_resumo_para_fonte(row: dict) -> dict:
    empresa = {
        "codigo": _texto_campo(row.get("codigo")),
        "razao_social": _texto_campo(row.get("empresa")),
        "cnpj": _texto_campo(row.get("cnpj")),
        "indice_lista": _texto_campo(row.get("indice_lista")),
        "linha_planilha": _texto_campo(row.get("linha_planilha")),
    }
    resultado = {
        "status": _texto_campo(row.get("status_execucao")),
        "motivo": _texto_campo(row.get("erro_resumo")),
        "erro_tipo": _texto_campo(row.get("erro_tipo")),
    }
    return {"empresa": empresa, "resultado": resultado}


def carregar_rows_resumo_existente(path_resumo: str):
    rows = []
    if not os.path.exists(path_resumo):
        return rows

    try:
        if path_resumo.lower().endswith(".xlsx"):
            try:
                from openpyxl import load_workbook
            except Exception as e:
                raise RuntimeError("Para ler resumo .xlsx instale a dependência openpyxl") from e

            wb = load_workbook(path_resumo, read_only=True, data_only=True)
            ws = wb.active
            linhas = ws.iter_rows(values_only=True)
            header_raw = next(linhas, None)
            if not header_raw:
                return rows
            header = [_texto_campo(valor) for valor in header_raw]
            for linha in linhas:
                row = {
                    header[i]: _texto_campo(valor)
                    for i, valor in enumerate(linha)
                    if i < len(header) and header[i]
                }
                rows.append(_registro_resumo_para_fonte(row))
        else:
            with open(path_resumo, "r", encoding="utf-8-sig", newline="") as f:
                reader = csv.DictReader(f, delimiter=";")
                for row in reader:
                    rows.append(_registro_resumo_para_fonte(row))
    except Exception:
        return rows

    return rows


def _caminhos_fonte_reprocessamento(output_base_dir: str, inicio=None, fim=None) -> list[str]:
    sufixo = _sufixo_lote(inicio, fim)
    caminhos = [
        os.path.join(output_base_dir, f"{RESUMO_BASE_NAME}{sufixo}.xlsx"),
        os.path.join(output_base_dir, f"{RESUMO_BASE_NAME}{sufixo}.csv"),
        os.path.join(output_base_dir, f"report_execucao_empresas{sufixo}.csv"),
    ]

    if inicio is not None and fim is not None:
        caminhos.extend(
            [
                os.path.join(output_base_dir, f"{RESUMO_BASE_NAME}.xlsx"),
                os.path.join(output_base_dir, f"{RESUMO_BASE_NAME}.csv"),
                os.path.join(output_base_dir, "report_execucao_empresas.csv"),
            ]
        )

    return caminhos


def carregar_registros_fonte_execucao(output_base_dir: str, inicio=None, fim=None):
    for caminho in _caminhos_fonte_reprocessamento(output_base_dir, inicio, fim):
        if not os.path.exists(caminho):
            continue
        try:
            if Path(caminho).name.startswith(RESUMO_BASE_NAME):
                registros = carregar_rows_resumo_existente(caminho)
            else:
                registros = carregar_rows_report_existente(caminho)
        except Exception as exc:
            print(f"Falha ao ler fonte de reprocessamento {caminho}: {type(exc).__name__}: {exc}")
            continue

        if registros:
            return registros, caminho

    return [], ""


def carregar_checkpoint(path_checkpoint: str):
    if not os.path.exists(path_checkpoint):
        return {"processadas": set(), "ultimo_indice": -1}

    try:
        with open(path_checkpoint, "r", encoding="utf-8") as f:
            data = json.load(f)
        processadas = set(re.sub(r"\D", "", c) for c in data.get("processadas", []) if c)
        ultimo_indice = int(data.get("ultimo_indice", -1))
        return {"processadas": processadas, "ultimo_indice": ultimo_indice}
    except Exception:
        return {"processadas": set(), "ultimo_indice": -1}


def salvar_checkpoint(path_checkpoint: str, processadas, ultimo_indice: int):
    payload = {
        "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "ultimo_indice": int(ultimo_indice),
        "processadas": sorted(set(processadas)),
    }
    tmp_path = f"{path_checkpoint}.tmp"
    with open(tmp_path, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)
    os.replace(tmp_path, path_checkpoint)


def executar_empresa(empresa: dict):
    inicio = datetime.now()
    ultimo_motivo = ""

    for tentativa in range(1, MAX_TENTATIVAS + 1):
        print("\n" + "=" * 80)
        print(f"Empresa {empresa['codigo']} - {empresa['razao_social']}")
        print(f"CNPJ: {empresa['cnpj']} | Segmento: {empresa['segmento']}")
        print("Senha prefeitura (referencia): [OCULTA]")
        print(f"Tentativa {tentativa}/{MAX_TENTATIVAS}")
        print(
            f"Etapa 1: login automatico (CNPJ/senha) + captcha humano; apos Entrar o robo navega sozinho para Lista Nota Fiscais. Tempo: {LOGIN_WAIT_SECONDS}s."
        )

        env = os.environ.copy()
        env["LOGIN_WAIT_SECONDS"] = str(LOGIN_WAIT_SECONDS)
        env["STRICT_LISTA_INICIAL"] = "1"
        env["EMPRESA_PASTA_FORCADA"] = nome_pasta_empresa_por_dados(empresa)
        env["AUTO_LOGIN_PREFEITURA"] = "1"
        env["EMPRESA_CNPJ"] = re.sub(r"\D", "", empresa["cnpj"])
        env["EMPRESA_SENHA"] = empresa["senha_prefeitura"]
        env["FORCAR_SESSAO_LIMPA_LOGIN"] = os.environ.get("FORCAR_SESSAO_LIMPA_LOGIN", "1")

        try:
            proc = executar_main_processo(env_extra=env, timeout=TIMEOUT_PROCESSO_MAIN)
        except FileNotFoundError as e:
            ultimo_motivo = str(e)
            print(f"Falha tentativa {tentativa}: {ultimo_motivo}")
            break
        except subprocess.TimeoutExpired:
            ultimo_motivo = f"Timeout de execucao do backend principal (> {TIMEOUT_PROCESSO_MAIN}s)"
            print(f"Falha tentativa {tentativa}: {ultimo_motivo}")
            continue

        rc = int(proc.returncode)
        sucesso_por_codigo = {
            0: ("SUCESSO", "OK"),
            EXIT_CODE_SEM_COMPETENCIA: ("SUCESSO_SEM_COMPETENCIA", "Sem notas na competencia alvo"),
            EXIT_CODE_SEM_SERVICOS: ("SUCESSO_SEM_SERVICOS", "Contribuinte sem modulo de Nota Fiscal"),
            EXIT_CODE_CREDENCIAL_INVALIDA: ("REVISAO_MANUAL", "Credencial invalida no portal"),
        }

        if rc == EXIT_CODE_TOMADOS_PDF_REVISAO_MANUAL:
            status_pdf = carregar_status_tomados_pdf_empresa(empresa)
            motivo = (status_pdf.get("codigo_erro") or "").strip() or "PDF_TOMADOS_NAO_GERADO"
            return {
                "status": "REVISAO_MANUAL",
                "motivo": motivo,
                "tentativas": tentativa,
                "acao_recomendada": inferir_acao_recomendada("REVISAO_MANUAL", motivo),
                "inicio": inicio,
                "fim": datetime.now(),
            }

        if rc == EXIT_CODE_EMPRESA_MULTIPLA:
            motivo = "EMPRESA_MULTIPLA"
            return {
                "status": "REVISAO_MANUAL",
                "motivo": motivo,
                "tentativas": tentativa,
                "acao_recomendada": inferir_acao_recomendada("REVISAO_MANUAL", motivo),
                "inicio": inicio,
                "fim": datetime.now(),
            }

        if rc == EXIT_CODE_APURACAO_COMPLETA_REVISAO_MANUAL:
            motivo = "APURACAO_COMPLETA_COM_FALHAS"
            return {
                "status": "REVISAO_MANUAL",
                "motivo": motivo,
                "tentativas": tentativa,
                "acao_recomendada": inferir_acao_recomendada("REVISAO_MANUAL", motivo),
                "inicio": inicio,
                "fim": datetime.now(),
            }

        if rc == EXIT_CODE_MULTI_CADASTRO_REVISAO_MANUAL:
            motivo = "MULTIPLOS_CADASTROS_COM_FALHAS"
            return {
                "status": "REVISAO_MANUAL",
                "motivo": motivo,
                "tentativas": tentativa,
                "acao_recomendada": inferir_acao_recomendada("REVISAO_MANUAL", motivo),
                "inicio": inicio,
                "fim": datetime.now(),
            }

        if rc in sucesso_por_codigo:
            status, motivo = sucesso_por_codigo[rc]
            return {
                "status": status,
                "motivo": motivo,
                "tentativas": tentativa,
                "acao_recomendada": inferir_acao_recomendada(status, motivo),
                "inicio": inicio,
                "fim": datetime.now(),
            }

        if rc == EXIT_CODE_CAPTCHA_TIMEOUT:
            ultimo_motivo = "Captcha nao resolvido a tempo"
        elif rc == EXIT_CODE_TOMADOS_FALHA:
            ultimo_motivo = "Falha na etapa de Servicos Tomados"
            print(f"Falha tentativa {tentativa}: {ultimo_motivo}")
            break
        elif rc == EXIT_CODE_CHROME_INIT_FALHA:
            ultimo_motivo = "Falha na inicializacao do Chrome/WebDriver"
            print(f"Falha tentativa {tentativa}: {ultimo_motivo}")
            break
        else:
            ultimo_motivo = f"Falha execucao (exit={rc})"

        print(f"Falha tentativa {tentativa}: {ultimo_motivo}")

    return {
        "status": "FALHA",
        "motivo": ultimo_motivo or "Falha desconhecida",
        "tentativas": MAX_TENTATIVAS,
        "acao_recomendada": inferir_acao_recomendada("FALHA", ultimo_motivo or "Falha desconhecida"),
        "inicio": inicio,
        "fim": datetime.now(),
    }

REPORT_HEADER = [
    "timestamp_inicio",
    "timestamp_fim",
    "codigo_empresa",
    "razao_social",
    "cnpj",
    "segmento",
    "status",
    "motivo",
    "tentativas",
    "acao_recomendada",
]


def garantir_report_com_header(path_report: str):
    if not os.path.exists(path_report):
        with open(path_report, "w", newline="", encoding="utf-8-sig") as f:
            csv.writer(f, delimiter=';').writerow(REPORT_HEADER)
        return

    with open(path_report, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f, delimiter=';')
        fieldnames = reader.fieldnames or []
        if fieldnames == REPORT_HEADER:
            return
        rows = list(reader)

    with open(path_report, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f, delimiter=';')
        writer.writerow(REPORT_HEADER)
        for row in rows:
            status = (row.get("status") or "").strip()
            motivo = (row.get("motivo") or "").strip()
            writer.writerow([
                (row.get("timestamp_inicio") or "").strip(),
                (row.get("timestamp_fim") or "").strip(),
                (row.get("codigo_empresa") or "").strip(),
                (row.get("razao_social") or "").strip(),
                (row.get("cnpj") or "").strip(),
                (row.get("segmento") or "").strip(),
                status,
                motivo,
                (row.get("tentativas") or "").strip(),
                (row.get("acao_recomendada") or "").strip() or inferir_acao_recomendada(status, motivo),
            ])


def resetar_report(path_report: str):
    with open(path_report, "w", newline="", encoding="utf-8-sig") as f:
        csv.writer(f, delimiter=';').writerow(REPORT_HEADER)


def append_report_row(path_report: str, row):
    garantir_report_com_header(path_report)
    with open(path_report, "a", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f, delimiter=';')
        w.writerow([
            row["inicio"].strftime("%Y-%m-%d %H:%M:%S"),
            row["fim"].strftime("%Y-%m-%d %H:%M:%S"),
            row["empresa"]["codigo"],
            row["empresa"]["razao_social"],
            row["empresa"]["cnpj"],
            row["empresa"]["segmento"],
            row["resultado"]["status"],
            row["resultado"]["motivo"],
            row["resultado"]["tentativas"],
            row["resultado"].get("acao_recomendada", ""),
        ])


def resolver_competencia_execucao() -> str:
    apuracao_ref = os.environ.get("APURACAO_REFERENCIA", "").strip() or datetime.now().strftime("%m/%Y")
    try:
        return competencia_alvo_dir_name(apuracao_ref)
    except Exception:
        return ""


def _normalizar_digitos(valor: str) -> str:
    return re.sub(r"\D", "", valor or "")


def _mapear_empresas_por_cnpj(empresas: list[dict]) -> dict[str, list[dict]]:
    mapa: dict[str, list[dict]] = {}
    for empresa in empresas:
        cnpj = _normalizar_digitos(empresa.get("cnpj", ""))
        if not cnpj:
            continue
        mapa.setdefault(cnpj, []).append(dict(empresa))
    return mapa


def _enriquecer_empresa_para_resumo(empresa: dict, mapa_empresas: dict[str, list[dict]]) -> dict:
    if empresa.get("indice_lista") and empresa.get("linha_planilha"):
        return dict(empresa)

    cnpj = _normalizar_digitos(empresa.get("cnpj", ""))
    candidatos = mapa_empresas.get(cnpj) or []
    if candidatos:
        base = dict(candidatos.pop(0))
        for chave, valor in empresa.items():
            if valor and not base.get(chave):
                base[chave] = valor
        return base

    return dict(empresa)


def _tem_arquivos_em_pasta(pasta: Path) -> bool:
    try:
        return pasta.exists() and any(pasta.iterdir())
    except Exception:
        return False


def _teve_prestados(artifacts) -> bool:
    try:
        if artifacts.log_downloads.exists():
            return True
        return _tem_arquivos_em_pasta(artifacts.competencia_dir / "PRESTADOS")
    except Exception:
        return False


def _teve_tomados(artifacts) -> bool:
    try:
        if artifacts.log_tomados.exists():
            return True
        return _tem_arquivos_em_pasta(artifacts.competencia_dir / "TOMADOS")
    except Exception:
        return False


def _teve_iss(artifacts) -> bool:
    try:
        if not artifacts.log_manual.exists():
            return False
        with open(artifacts.log_manual, "r", encoding="utf-8") as f:
            for line in f:
                upper = line.upper()
                if "GUIA_ISS=" in upper and "SKIP_PERFIL" not in upper:
                    return True
    except Exception:
        pass

    try:
        for padrao in ("PRESTADOS/GUIA_ISS_*.pdf", "TOMADOS/GUIA_ISS_*.pdf"):
            if list(artifacts.competencia_dir.glob(padrao)):
                return True
    except Exception:
        pass

    return False


def _resumir_texto(texto: str, limite: int = 180) -> str:
    texto_limpo = re.sub(r"\s+", " ", (texto or "")).strip()
    return texto_limpo[:limite]


def _classificar_erro_execucao(status: str, motivo: str, erro_tipo_previo: str = "") -> tuple[str, str]:
    erro_tipo = classificar_erro_execucao(status, motivo, erro_tipo_previo)
    erro_resumo = _resumir_texto(motivo, 180) if erro_tipo else ""
    return erro_tipo, erro_resumo


def construir_linhas_resumo_execucao(
    resultados: list[dict],
    empresas: list[dict],
    output_base_dir: str,
    competencia_dir_name: str = "",
) -> list[dict]:
    runtime_paths = build_runtime_paths(Path(get_runtime_base()), Path(output_base_dir))
    competencia_dir_name = (competencia_dir_name or "").strip() or resolver_competencia_execucao()
    mapa_empresas = _mapear_empresas_por_cnpj(empresas)
    linhas: list[dict] = []

    for row in resultados:
        empresa_base = _enriquecer_empresa_para_resumo(row.get("empresa") or {}, mapa_empresas)
        resultado = row.get("resultado") or {}
        artifacts = ARTIFACT_LOCATOR.get_company_artifacts(runtime_paths, empresa_base, competencia_dir_name=competencia_dir_name)
        status_execucao = (resultado.get("status") or "").strip()
        erro_tipo, erro_resumo = _classificar_erro_execucao(
            status_execucao,
            resultado.get("motivo") or "",
            resultado.get("erro_tipo") or "",
        )
        competencia = artifacts.competencia_dir.name if re.fullmatch(r"\d{2}\.\d{4}", artifacts.competencia_dir.name or "") else ""

        linhas.append(
            {
                "indice_lista": empresa_base.get("indice_lista", ""),
                "linha_planilha": empresa_base.get("linha_planilha", ""),
                "empresa": empresa_base.get("razao_social", ""),
                "codigo": empresa_base.get("codigo", ""),
                "cnpj": empresa_base.get("cnpj", ""),
                "competencia": competencia,
                "status_execucao": status_execucao,
                "teve_iss": "SIM" if _teve_iss(artifacts) else "NAO",
                "teve_prestados": "SIM" if _teve_prestados(artifacts) else "NAO",
                "teve_tomados": "SIM" if _teve_tomados(artifacts) else "NAO",
                "erro_tipo": erro_tipo,
                "erro_resumo": erro_resumo,
            }
        )

    return linhas


def escrever_resumo_execucao(path_resumo: str, linhas: list[dict]) -> str:
    if path_resumo.lower().endswith(".xlsx"):
        if Workbook is None:
            raise RuntimeError("openpyxl nao disponivel para gerar resumo XLSX.")

        wb = Workbook()
        ws = wb.active
        ws.title = "Resumo"
        ws.append(RESUMO_HEADER)
        for linha in linhas:
            ws.append([linha.get(col, "") for col in RESUMO_HEADER])
        if ws.max_row > 1:
            ws.auto_filter.ref = ws.dimensions
        ws.freeze_panes = "A2"
        wb.save(path_resumo)
        return path_resumo

    with open(path_resumo, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f, delimiter=";")
        writer.writerow(RESUMO_HEADER)
        for linha in linhas:
            writer.writerow([
                linha.get("indice_lista", ""),
                linha.get("linha_planilha", ""),
                linha.get("empresa", ""),
                linha.get("codigo", ""),
                linha.get("cnpj", ""),
                linha.get("competencia", ""),
                linha.get("status_execucao", ""),
                linha.get("teve_iss", ""),
                linha.get("teve_prestados", ""),
                linha.get("teve_tomados", ""),
                linha.get("erro_tipo", ""),
                linha.get("erro_resumo", ""),
            ])
    return path_resumo


def salvar_resumo_execucao(
    resultados: list[dict],
    empresas: list[dict],
    output_base_dir: str,
    inicio=None,
    fim=None,
    competencia_dir_name: str = "",
) -> str | None:
    linhas = construir_linhas_resumo_execucao(resultados, empresas, output_base_dir, competencia_dir_name=competencia_dir_name)
    path_resumo = resolver_path_resumo_execucao(output_base_dir, inicio, fim)

    try:
        return escrever_resumo_execucao(path_resumo, linhas)
    except Exception as exc:
        print(f"Falha ao gerar resumo em {path_resumo}: {type(exc).__name__}: {exc}")
        if path_resumo.lower().endswith(".xlsx"):
            fallback = str(Path(path_resumo).with_suffix(".csv"))
            try:
                print(f"Tentando fallback CSV: {fallback}")
                return escrever_resumo_execucao(fallback, linhas)
            except Exception as fallback_exc:
                print(f"Falha ao gerar resumo CSV: {type(fallback_exc).__name__}: {fallback_exc}")
        return None


def main():
    if not os.path.exists(CSV_EMPRESAS):
        print(f"Arquivo de empresas nÃ£o encontrado: {CSV_EMPRESAS}")
        print("Informe EMPRESAS_ARQUIVO com caminho vÃ¡lido (.xlsx/.csv).")
        return

    empresas = carregar_empresas(CSV_EMPRESAS)
    print(f"Empresas carregadas: {len(empresas)}")

    try:
        faixa_inicio, faixa_fim, empresas = resolver_faixa_execucao(empresas)
    except ValueError as e:
        raise SystemExit(str(e)) from e
    report_path, checkpoint_path = resolver_paths_execucao(OUTPUT_BASE_DIR, faixa_inicio, faixa_fim)
    competencia_dir_name = resolver_competencia_execucao()
    try:
        filtros = resolver_filtros_execucao()
    except ValueError as e:
        raise SystemExit(str(e)) from e
    filtros_ativos = bool(filtros["empresas"] or filtros["erro_tipos"])
    usar_retomada = CONTINUAR_DE_ONDE_PAROU and not filtros_ativos
    usar_checkpoint_retomada = USAR_CHECKPOINT and not filtros_ativos
    registros_fonte = []
    caminho_fonte = ""

    if filtros_ativos and CONTINUAR_DE_ONDE_PAROU:
        print("Filtros seletivos ativos: a retomada por report/checkpoint antigo sera ignorada nesta execucao.")
    if filtros["erro_tipos"]:
        registros_fonte, caminho_fonte = carregar_registros_fonte_execucao(OUTPUT_BASE_DIR, faixa_inicio, faixa_fim)
        if not registros_fonte:
            raise SystemExit(
                "FILTRAR_ERRO_TIPO foi informado, mas nenhum resumo/report de origem foi encontrado para o escopo atual."
            )
    empresas = filtrar_empresas_por_criterios(
        empresas,
        filtros["empresas"],
        filtros["erro_tipos"],
        registros_fonte if filtros["erro_tipos"] else None,
    )

    faixa_txt = "todas" if faixa_inicio is None else f"{faixa_inicio}-{faixa_fim}"
    print(f"Faixa aplicada: {faixa_txt}")
    print(f"Empresas selecionadas: {len(empresas)}")
    print(f"Report: {report_path}")
    print(f"Checkpoint: {checkpoint_path}")
    print(f"Filtro empresas: {', '.join(filtros['empresas']) if filtros['empresas'] else 'nenhum'}")
    print(f"Filtro erro: {', '.join(filtros['erro_tipos']) if filtros['erro_tipos'] else 'nenhum'}")
    if caminho_fonte:
        print(f"Fonte de erro: {caminho_fonte}")
    if not empresas:
        print("Nenhuma empresa foi selecionada pelos filtros atuais. Nada a executar.")
        return

    concluidas = set()
    if usar_retomada:
        concluidas = carregar_report_existente(report_path)
        if concluidas:
            print(f"Retomada ativa: {len(concluidas)} empresas jÃ¡ concluÃ­das serÃ£o puladas.")

    start_idx = 0
    processadas_checkpoint = set()
    ck_ultimo_indice = -1
    if usar_checkpoint_retomada:
        ck = carregar_checkpoint(checkpoint_path)
        processadas_checkpoint = ck["processadas"]
        ck_ultimo_indice = int(ck.get("ultimo_indice", -1))
        if processadas_checkpoint:
            print(f"Checkpoint ativo: {len(processadas_checkpoint)} empresas já processadas serão puladas.")
        if usar_retomada and ck_ultimo_indice >= 0:
            start_idx = ck_ultimo_indice + 1
            print(f"Retomando a partir do índice {start_idx} (checkpoint).")

    ja_processadas = set(concluidas) | set(processadas_checkpoint)

    resultados = carregar_rows_report_existente(report_path) if usar_retomada else []
    if usar_retomada:
        garantir_report_com_header(report_path)
    else:
        resetar_report(report_path)
    cnpjs_ja_no_report = {re.sub(r"\D", "", r["empresa"].get("cnpj", "")) for r in resultados if r.get("empresa")}

    last_idx_concluido = start_idx - 1
    try:
        for idx, empresa in enumerate(empresas[start_idx:], start=start_idx):
            cnpj_limpo = re.sub(r"\D", "", empresa.get("cnpj", ""))

            if cnpj_limpo and cnpj_limpo in ja_processadas:
                # Atualiza checkpoint mesmo quando pula, para retomar do ponto exato
                last_idx_concluido = idx
                if USAR_CHECKPOINT:
                    salvar_checkpoint(checkpoint_path, processadas_checkpoint, idx)

                if cnpj_limpo in cnpjs_ja_no_report:
                    continue

                agora = datetime.now()
                resultado = {
                    "status": "SUCESSO",
                    "motivo": "Pulada por retomada/checkpoint (já processada)",
                    "tentativas": 0,
                    "acao_recomendada": "",
                    "inicio": agora,
                    "fim": agora,
                }
                row = {"empresa": empresa, "resultado": resultado, "inicio": agora, "fim": agora}
                resultados.append(row)
                cnpjs_ja_no_report.add(cnpj_limpo)
                append_report_row(report_path, row)
                continue

            try:
                res = executar_empresa(empresa)
            except Exception as e:
                agora = datetime.now()
                res = {
                    "status": "FALHA",
                    "motivo": f"Erro inesperado no orquestrador: {str(e)[:160]}",
                    "tentativas": 0,
                    "acao_recomendada": inferir_acao_recomendada(
                        "FALHA",
                        f"Erro inesperado no orquestrador: {str(e)[:160]}",
                    ),
                    "inicio": agora,
                    "fim": agora,
                }

            if not res.get("acao_recomendada"):
                res["acao_recomendada"] = inferir_acao_recomendada(
                    res.get("status", ""),
                    res.get("motivo", ""),
                )

            row = {"empresa": empresa, "resultado": res, "inicio": res["inicio"], "fim": res["fim"]}
            resultados.append(row)
            if cnpj_limpo:
                cnpjs_ja_no_report.add(cnpj_limpo)
            append_report_row(report_path, row)

            if cnpj_limpo and res.get("status") in STATUS_FINAIS_NAO_REPROCESSAR:
                ja_processadas.add(cnpj_limpo)
                if USAR_CHECKPOINT:
                    processadas_checkpoint.add(cnpj_limpo)

            last_idx_concluido = idx
            if USAR_CHECKPOINT:
                # Atualiza SEMPRE o índice (sucesso ou falha) para retomar do ponto exato
                salvar_checkpoint(checkpoint_path, processadas_checkpoint, idx)

    except KeyboardInterrupt:
        print("\nInterrompido pelo usuário (Ctrl+C). Salvando checkpoint e encerrando...")
        if USAR_CHECKPOINT:
            salvar_checkpoint(checkpoint_path, processadas_checkpoint, last_idx_concluido)
        return
    except Exception:
        print("\nErro inesperado no loop principal. Salvando checkpoint antes de sair...")
        print(traceback.format_exc())
        if USAR_CHECKPOINT:
            salvar_checkpoint(checkpoint_path, processadas_checkpoint, last_idx_concluido)
        raise

    total = len(resultados)
    ok = sum(1 for r in resultados if r["resultado"]["status"] in STATUS_SUCESSO)
    revisao_manual = sum(1 for r in resultados if r["resultado"]["status"] == "REVISAO_MANUAL")
    falha = sum(1 for r in resultados if r["resultado"]["status"] not in STATUS_FINAIS_NAO_REPROCESSAR)
    print("\n" + "=" * 80)
    print(f"Processamento finalizado. Total={total} | Sucesso={ok} | RevisaoManual={revisao_manual} | Falha={falha}")
    print(f"Report: {report_path}")

    if USAR_CHECKPOINT and os.path.exists(checkpoint_path):
        if falha == 0:
            os.remove(checkpoint_path)
            print(f"Checkpoint removido ao final: {checkpoint_path}")
        else:
            print(f"Checkpoint mantido para retomada: {checkpoint_path}")

    try:
        resumo_path = salvar_resumo_execucao(
            resultados,
            empresas,
            OUTPUT_BASE_DIR,
            faixa_inicio,
            faixa_fim,
            competencia_dir_name=competencia_dir_name,
        )
        if resumo_path:
            print(f"Resumo: {resumo_path}")
    except Exception:
        print("Falha inesperada ao consolidar o resumo geral.")
        print(traceback.format_exc())


if __name__ == "__main__":
    main()
