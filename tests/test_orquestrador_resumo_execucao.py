from pathlib import Path

import orquestrador_empresas as oq


def _empresa(codigo: str, cnpj: str, indice_lista: int, linha_planilha: int, razao_social: str):
    return {
        "codigo": codigo,
        "razao_social": razao_social,
        "cnpj": cnpj,
        "segmento": "SERVICOS",
        "senha_prefeitura": "segredo",
        "indice_lista": indice_lista,
        "linha_planilha": linha_planilha,
    }


def _preparar_artifacts(tmp_path, empresa, competencia="01.2026", com_logs=True):
    output_dir = tmp_path / "output"
    runtime_paths = oq.build_runtime_paths(tmp_path / "project", output_dir)
    pasta_empresa = runtime_paths.downloads_dir / oq.nome_pasta_empresa_por_dados(empresa)
    geral_dir = pasta_empresa / competencia / "_GERAL"
    geral_dir.mkdir(parents=True, exist_ok=True)

    if com_logs:
        (geral_dir / "log_downloads_nfse.txt").write_text("download", encoding="utf-8")
        (geral_dir / "log_tomados.txt").write_text("tomados", encoding="utf-8")
        (geral_dir / "log_fechamento_manual.txt").write_text("GUIA_ISS=1\n", encoding="utf-8")

    return output_dir


def _ler_resumo(path_resumo: Path):
    if path_resumo.suffix.lower() == ".xlsx":
        from openpyxl import load_workbook

        wb = load_workbook(path_resumo, read_only=True, data_only=True)
        ws = wb.active
        linhas = list(ws.iter_rows(values_only=True))
        header = ["" if valor is None else str(valor) for valor in linhas[0]]
        rows = [
            {header[idx]: "" if valor is None else str(valor) for idx, valor in enumerate(linha)}
            for linha in linhas[1:]
        ]
        return header, rows

    import csv

    with open(path_resumo, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f, delimiter=";")
        header = list(reader.fieldnames or [])
        rows = [{k: "" if v is None else str(v) for k, v in row.items()} for row in reader]
        return header, rows


def test_salvar_resumo_execucao_gera_planilha_com_colunas_minimas_sem_faixa(tmp_path):
    empresa_ok = _empresa("1", "12.345.678/0001-90", 1, 2, "Empresa OK")
    empresa_falha = _empresa("2", "98.765.432/0001-10", 2, 3, "Empresa Falha")
    output_dir = _preparar_artifacts(tmp_path, empresa_ok, com_logs=True)
    _preparar_artifacts(tmp_path, empresa_falha, com_logs=False)

    resultados = [
        {
            "empresa": empresa_ok,
            "resultado": {
                "status": "SUCESSO",
                "motivo": "OK",
                "tentativas": 1,
                "acao_recomendada": "",
            },
            "inicio": oq.datetime.now(),
            "fim": oq.datetime.now(),
        },
        {
            "empresa": empresa_falha,
            "resultado": {
                "status": "FALHA",
                "motivo": "Timeout na etapa final",
                "tentativas": 3,
                "acao_recomendada": "",
            },
            "inicio": oq.datetime.now(),
            "fim": oq.datetime.now(),
        },
    ]

    path_resumo = Path(
        oq.salvar_resumo_execucao(
            resultados,
            [empresa_ok, empresa_falha],
            str(output_dir),
            competencia_dir_name="01.2026",
        )
    )

    assert path_resumo.exists()
    ext_esperada = ".xlsx" if oq.RESUMO_USA_XLSX else ".csv"
    assert path_resumo.name == f"resumo_execucao_empresas{ext_esperada}"

    header, rows = _ler_resumo(path_resumo)
    assert header == oq.RESUMO_HEADER
    assert len(rows) == 2

    primeira = rows[0]
    assert primeira["indice_lista"] == "1"
    assert primeira["linha_planilha"] == "2"
    assert primeira["empresa"] == "Empresa OK"
    assert primeira["codigo"] == "1"
    assert primeira["cnpj"] == "12.345.678/0001-90"
    assert primeira["competencia"] == "01.2026"
    assert primeira["status_execucao"] == "SUCESSO"
    assert primeira["teve_iss"] == "SIM"
    assert primeira["teve_prestados"] == "SIM"
    assert primeira["teve_tomados"] == "SIM"
    assert primeira["erro_tipo"] == ""
    assert primeira["erro_resumo"] == ""

    segunda = rows[1]
    assert segunda["indice_lista"] == "2"
    assert segunda["linha_planilha"] == "3"
    assert segunda["empresa"] == "Empresa Falha"
    assert segunda["status_execucao"] == "FALHA"
    assert segunda["teve_iss"] == "NAO"
    assert segunda["teve_prestados"] == "NAO"
    assert segunda["teve_tomados"] == "NAO"
    assert segunda["erro_tipo"] == "TIMEOUT"
    assert segunda["erro_resumo"] == "Timeout na etapa final"


def test_salvar_resumo_execucao_nomeia_arquivo_por_lote(tmp_path):
    empresa = _empresa("1", "12.345.678/0001-90", 1, 2, "Empresa Lote")
    output_dir = _preparar_artifacts(tmp_path, empresa, com_logs=True)
    resultados = [
        {
            "empresa": empresa,
            "resultado": {
                "status": "SUCESSO",
                "motivo": "OK",
                "tentativas": 1,
                "acao_recomendada": "",
            },
            "inicio": oq.datetime.now(),
            "fim": oq.datetime.now(),
        }
    ]

    path_resumo = Path(
        oq.salvar_resumo_execucao(
            resultados,
            [empresa],
            str(output_dir),
            inicio=1,
            fim=100,
            competencia_dir_name="01.2026",
        )
    )

    ext_esperada = ".xlsx" if oq.RESUMO_USA_XLSX else ".csv"
    assert path_resumo.name == f"resumo_execucao_empresas__lote_001_100{ext_esperada}"
